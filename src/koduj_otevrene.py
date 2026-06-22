# -*- coding: utf-8 -*-
"""Respekt - kodovani 4 otevrenych otazek (153, 154, 155, 156).
Pravidlova multi-label klasifikace zakotvena ve cteni vzorku. Per-otazku,
valence oddelena. col 154 = extrakce jmenovanych zdroju (entit), ne tematu.
Vystup: xlsx workbook + strojove citelne kodovani (CSV)."""
import pandas as pd, numpy as np, unicodedata, re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = ROOT / 'data' / 'raw'
OUT_DIR = ROOT / 'data' / 'processed'
OUT_DIR.mkdir(parents=True, exist_ok=True)

def newest_export():
    files = sorted(RAW_DIR.glob('SurveyHeroResponses*.csv'))
    if not files:
        raise SystemExit(f'Zadny export v {RAW_DIR}/ (ocekavam SurveyHeroResponses*.csv)')
    return files[-1]

SRC = newest_export()
print(f'Vstup: {SRC.name}')
df = pd.read_csv(SRC, dtype=str, keep_default_na=False)
df = df[df['Status'] == 'Completed'].reset_index(drop=True)

def norm(s):
    s = str(s).strip().lower()
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    return re.sub(r'\s+', ' ', s)

def clean(s):
    s = re.sub(r'\s+', ' ', str(s)).strip()
    if s[:1] in '=+-@':
        s = ' ' + s
    return s[:160]

# ====================== TAXONOMIE (155, 156, 153) ======================
TAX = {
 '155': {  # V cem vyjimecny - CHVALA (pozitivni valence)
  'Hloubka, kontext, souvislosti': r'hloubk|do hloubky|kontext|souvislost|analyz|porozumen|pricin|vysvetl|rozbor|nadhled|odstup',
  'Sire a vyber temat': r'temat|\btema\b|sirk|\bsire\b|zaber|okruh|pestrost|rozmanit|rozkroc|spektrum|zajimav tema',
  'Kvalita psani, jazyk, uroven': r'kvalit|jazyk|\bstyl|psani|cestin|gramat|literar|napsan|uroven|profesional|preciz',
  'Autori, osobnosti, redakce': r'autor|osobnost|redaktor|redakc|novinar|pisatel|kundra|spurny|erudic|\btym\b',
  'Nezavislost': r'nezavisl|bez vlivu|\bsvobod|oligarch|nepatri|neslouzi',
  # byvale siroke "Duvera, verohodnost, vyvazenost" (325) rozdeleno na 2 (TODO 1a):
  'Duvera, verohodnost, fakta': r'duver|verohodn|overen|\bfakt|pravd|serioz|presnost|spolehliv',
  'Vyvazenost, objektivita, nestrannost': r'vyvazen|objektiv|nestrann|obe strany|vsechny strany|ruzne (?:uhly|pohled|nazor)|vsechny uhly',
  'Investigativa, reportaze': r'investigat|reportaz|\bterenu|na miste|odhal|\bkauz',
  'Hodnotove souzneni, svetonazor': r'liberal|\bhodnot|svetonazor|vnimani sveta|videni sveta|ztotoznuj|souzn|progresiv|stejny pohled|naladen',
  # z "Formy" vytazeny preklady/zahranicni obsah jako samostatne tema (TODO 1a):
  'Preklady, zahranicni obsah': r'preklad|zahranicni\w* (?:media|clank|texty|zpravodaj)|ze zahranic\w* medii|prejima\w* (?:ze zahranic|clanky)|spoluprac\w*.{0,22}(?:atlantic|nyt|economist|guardian|new yorker|times)|vyber prekladu',
  'Forma (bez reklam, audio, epub)': r'bez reklam|\baudio|namluv|mezi novinami|\bepub|\bformat|papirov\w* (?:verze|forma|vydani)',
  'Nepodbizivost, odvaha': r'nepodbiz|odvah|odvazn|kritick|\btabu|jde proti|neboji|provokativ',
  'Tradice, znacka': r'tradic|dlouholet|\bznacka|historie|leta ctu|pres \d+ let',
 },
 '156': {  # Co vadi / zlepsit - VYTKY/NAMETY. Siroka temata Web/app a Obsah rozdelena (TODO 1a/1c).
  'Nic nevadi / spokojenost': r'^nic\b|nic mi nevad|nevadi nic|jsem spokoj|jsme spokoj|\bvse ok\b|nenapada|nemam co vyt|nemam vyhrad|nemam pripomink|\bhappy\b|^nevim$|^nic$',
  # --- byvale siroke "Web/aplikace/UX/technika" rozdeleno na 4 (vyhledavani a prehlednost uz byly samostatne) ---
  'Web / navigace': r'\bweb|listovat|posunovat|mezi clanky|na webu|prochazet|navigac',
  'Aplikace / technika (UX, bugy, vykon)': (
      r'aplikac|\bapp\b|appka|rozhrani|uzivatelsk|ovlada|zamrz(?!el)|zaseka|sekani|\bseka\b'
      r'|spadl|spadne|nefunguj|nestabil|pomal|kde jsem (?:skoncil|prestal)|vybij\w* bater|spotreb\w* dat'
      r'|technick\w* (?:probl|aspekt|zalezitost|chyb|vec|stranka|podmink|reseni|stav|zavad)',
      r'vyhledav|prehledn|\barchiv|filtrov|hledat star'),
  'Vyhledavani': r'vyhledav|hledat (?:clanek|autor|stare)|filtrov|dle autor',
  'Prehlednost / archiv / orientace': r'prehledn|\barchiv|razeni|orientac|struktur|hledat star',
  'Audio / AI hlas (smisene)': r'\baudio|namluv|\bhlas|umel|\bai\b|nadech|robot|chybne cteni|audioverz',
  # --- byvale siroke "Obsah / chybejici temata" rozdeleno na 3 ---
  'Kulturni rubrika': r'kultur|recenz|\btipy',
  'Chybejici oblasti': r'ekonom|\bsport|region|\bveda|domaci tema',
  'Ton / vice pozitivniho': r'pozitivn|nadej|depres|optimis|beznadej',
  'Grafika / obalka / ilustrace': r'grafik|obalk|titulni stran|reisenauer|raisenauer|riesenauer|ilustrac|obrazek|kreslen|kresb',
  'Jednostrannost / bias / aktivismus': r'jednostr|\blevic|liberal|aktivis|\bwoke|zaujat|ideolog|propag|gender|feminist|korektn|arogan|protistran|naladeni|tendencn',
  'Delka clanku (moc dlouhe)': r'\bdlouh|\bdelka|zkrat|kratsi|strucnej|rozvlacn|moc textu',
  'Cetnost / vice obsahu, podcastu': r'cetnost|casteji|jen tyden|aktualizac|frekvenc|\bdenne|vic podcast|malo podcast|vic clank|malo clank|kazdodenn',
  'Doruceni tisku': r'doruc|nechodi|\bpozde|schrank|\bposta|\bvcas',
  'Cena / paywall': r'\bcena|\bdrah|\bdraz|paywall|levnej|\bstoji|penez',
  'Reklamy': r'reklam',
 },
 '153': {  # Poslech v aplikaci - REFRAME: hl. proc jina platforma + problemy
  'Nezkousel / neposloucha v app': r'nezkous|nepouzivam aplikac|neposloucham|nevyuzivam|jen ctu|neslysel|^ne[.,]?$|^nezkousel|^nezkousela|nepreferuj',
  'Jina platforma / zvyklost': r'zvykl|\bzvyk|spotify|apple|youtube|podcast addict|pocket cast|antennapod|jine prostredi|jiny prehravac|jina aplikace|obecn\w* aplikac|preferuji web',
  'Duvod: vse na jednom miste (agregace)': r'na jednom miste|vsechny podcasty|vsechny na jednom|ruzn\w* zdroj|ruzn\w* podcasty|do fronty|\bfronta|i (?:ty )?jine|vidim nabidku|nabidku na jednom',
  'Duvod: sledovani prehraneho': r'co jsem (?:uz )?slysel|uz slysel|prehran|oznacen|navazuje|kde jsem skoncil|zaznamen',
  # Pozitivni hodnoceni MUSI byt o aplikaci Respektu: lookahead na pozitivni jadro
  # (uzke - bez holeho "vyhovuje"/"spokojen", ktere chytalo i chvalu Spotify/Apple)
  # + negativni lookahead vyrazujici negace (nevyhovuje, nebyl/nejsem spokojen, prestal).
  'Funguje dobre / spokojen':
      r'^(?!.*(?:nevyhovuj|nespokojen|nebyl\w*[\s\w]{0,8}spokojen|nejsem spokojen|prestal))'
      r'(?=.*(?:funguje (?:to |mi )?(?:dobre|skvele|bezvadne|v pohode)|je to super|je to ok'
      r'|je to (?:fajn|dobre|v pohode)|bez problemu?|jsem (?:velmi |maximalne )?spokojen'
      r'|v podstate mi vyhovuje|vyhovuje mi (?:v aplikaci|aplikace|appka|poslech)))',
  'Vytka: ovladani / prehlednost': r'prehledn|ovlada|rozhrani|navigac|tlacit|nemuzu najit|neprehledn',
  'CarPlay / Auto / Bluetooth': r'carplay|android auto|bluetooth|v aute|reproduk',
  'Technicky problem (prehravani/pozice)': r'\bnelze|\bnejde|nefunguj|\bchyba|nenacita|nespusti|spadne|zasekne|\bposun|\bskace|preskak|vraci|od zacatku|na pozadi|zamc|vypne|zastav|prerus|ztrac|nepamatuje',
  'Kvalita AI hlasu': r'\bhlas|umel|\bai\b|\brobot|monoton|prizvuk|vyslov|cte spatne',
  'Posloucha jen audioverze clanku': r'audioverz|prepis clanku|nactene clanky|audio clank|nestiham cist|nestihnu (?:precist|cist)',
 },
}
QCOL = {'155': 155, '156': 156, '153': 153}

# ====================== col 154 - JMENOVANE ZDROJE ======================
# (display, typ, pattern)
OUTLETS = [
 ('Denik N', 'Domaci zpravodajstvi', r'denik ?n\b|deniku ?n|dennik n\b|denik\.n'),
 ('Seznam Zpravy / Seznam.cz', 'Domaci zpravodajstvi', r'seznam'),
 ('Novinky.cz', 'Domaci zpravodajstvi', r'novinky'),
 ('Aktualne.cz', 'Domaci zpravodajstvi', r'aktualne'),
 ('Hospodarske noviny', 'Domaci zpravodajstvi', r'hospodarsk|\bhn\b|ihned'),
 ('iDnes / MF Dnes', 'Domaci zpravodajstvi', r'idnes|mf ?dnes|mlada fronta'),
 ('CT / CT24', 'Verejnopravni (domaci)', r'ceska televiz|\bct ?24|\bct1\b|168 hodin|\bct\b'),
 ('iRozhlas / Cesky rozhlas', 'Verejnopravni (domaci)', r'irozhlas|cesky rozhlas|\bcro\b|radiozurnal|\brozhlas'),
 ('CTK', 'Verejnopravni (domaci)', r'\bctk\b|ceska tiskova'),
 ('Lidovky / Lidove noviny', 'Domaci tisk / komentar', r'lidov'),
 ('Pravo', 'Domaci tisk / komentar', r'\bpravo\b'),
 ('Echo', 'Domaci tisk / komentar', r'\becho'),
 ('Reflex', 'Domaci tisk / komentar', r'reflex'),
 ('Forum24', 'Alternativni / investigativni', r'forum ?24'),
 ('A2larm / A2', 'Alternativni / investigativni', r'a2larm|\ba2\b'),
 ('Denik Referendum', 'Alternativni / investigativni', r'referendum'),
 ('Hlidaci pes / Investigace / Neovlivni / Voxpot', 'Alternativni / investigativni', r'hlidaci|investigace|neovlivni|voxpot'),
 ('DVTV', 'Influencer / osobnost / video', r'\bdvtv'),
 ('New York Times', 'Zahranicni', r'new york times|\bnyt\b'),
 ('The Economist', 'Zahranicni', r'economist'),
 ('Guardian', 'Zahranicni', r'guardian'),
 ('BBC', 'Zahranicni', r'\bbbc\b'),
 ('Dalsi zahranicni (FT, WaPo, Politico, Spiegel...)', 'Zahranicni', r'financial times|\bft\b|washington post|wall street|\bwsj\b|politico|atlantic|spiegel|le monde|reuters|bloomberg|al jazeera|\bnpr\b|foreign affairs|die zeit|sueddeutsch'),
 ('Slovenske (SME, Dennik N, Aktuality)', 'Slovenske', r'\bsme\b|aktuality|dennik n|denník'),
 ('Podcast (obecne i jmenovite)', 'Podcast', r'podcast|vinohradsk|studio n|brifink|kecy a politik|kecy & politik|dobrovsky|sidlo|5:59|vlevo dole|stardust|stastne pondeli'),
 ('YouTube', 'Socialni site / video', r'youtube'),
 ('Socialni site (FB, X, Instagram, Bluesky, TikTok)', 'Socialni site / video', r'facebook|\bfb\b|twitter|instagram|bluesky|tiktok|socialni sit'),
 ('Newsletter / Substack', 'Newsletter', r'newsletter|substack'),
]

# ====================== KODOVANI ======================
coded = pd.DataFrame({'ID': df['ID'].values})
results = {}

def slug(s):
    s = norm(s); s = re.sub(r'[^a-z0-9]+', '_', s); return s.strip('_')[:28]

for qkey, themes in TAX.items():
    ci = QCOL[qkey]; raw = df.iloc[:, ci]; nv = raw.map(norm)
    has = raw.str.strip() != ''; Nt = int(has.sum())
    rows = []
    for theme, pat in themes.items():
        # pat muze byt retezec, nebo dvojice (include, exclude) pro MECE odecet
        # specifictejsiho tematu (princip de-dup: sirsi kos nepocita to, co ma
        # vlastni uzsi tema – napr. Aplikace/technika neobsahuje Vyhledavani).
        if isinstance(pat, tuple):
            inc, exc = pat
            m = nv.str.contains(re.compile(inc)) & ~nv.str.contains(re.compile(exc)) & has
        else:
            m = nv.str.contains(re.compile(pat)) & has
        idx = list(raw.index[m]); n = len(idx)
        ex = [clean(raw.iloc[i]) for i in idx[:2]]
        coded[f'{qkey}_{slug(theme)}'] = m.values
        rows.append({'tema': theme, 'n': n, 'ex': ex})
    rows.sort(key=lambda r: -r['n'])
    results[qkey] = {'N': Nt, 'rows': rows}

# 154
ci = 154; raw = df.iloc[:, ci]; nv = raw.map(norm); has = raw.str.strip() != ''; N154 = int(has.sum())
out_rows = []; type_counts = {}
type_masks = {}
for disp, typ, pat in OUTLETS:
    m = nv.str.contains(re.compile(pat)) & has
    idx = list(raw.index[m]); n = len(idx)
    coded[f'154_src_{slug(disp)}'] = m.values
    out_rows.append({'zdroj': disp, 'typ': typ, 'n': n})
    type_masks.setdefault(typ, pd.Series(False, index=df.index))
    type_masks[typ] = type_masks[typ] | m
out_rows.sort(key=lambda r: -r['n'])
for typ, mask in type_masks.items():
    coded[f'154_typ_{slug(typ)}'] = mask.values
    type_counts[typ] = int(mask.sum())
type_rows = sorted(type_counts.items(), key=lambda kv: -kv[1])

coded.to_csv(OUT_DIR / 'respekt_otevrene_kodovano.csv', index=False)

# ====================== XLSX ======================
RED='CE0B24'; INK='181818'; LGREY='ECEFF2'; MGREY='868D98'
F = lambda **k: Font(name='Arial', **k)
thin = Side(style='thin', color='D9DEE3')
BORD = Border(bottom=thin)

wb = Workbook()

def style_header(ws, row, ncol):
    for c in range(1, ncol+1):
        cell = ws.cell(row=row, column=c)
        cell.font = F(bold=True, color='FFFFFF', size=10)
        cell.fill = PatternFill('solid', fgColor=INK)
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

def title_block(ws, title, sub):
    ws['A1'] = title; ws['A1'].font = F(bold=True, size=14, color=RED)
    ws['A2'] = sub; ws['A2'].font = F(size=10, color=MGREY); ws['A2'].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 30

# --- Prehled ---
ws = wb.active; ws.title = 'Prehled'
title_block(ws, 'Respekt - kodovani otevrenych otazek',
            'Pruzkum 2026-06-20, n=2139 dokoncenych. Ctyri otevrene otazky: per-otazku taxonomie, valence oddelena, col 154 jako jmenovane zdroje.')
notes = [
 '',
 'Metoda: pravidlova multi-label klasifikace (klicova slova), zakotvena ve cteni vzorku. Jedna odpoved muze nest vice tematu.',
 'Cetnosti jsou priblizne (zachycuji jasne formulace; nuance/preklepy mohou unikat). Priklady jsou doslovne (zkracene).',
 'Valence je oddelena: chvalu (155) a vytky (156) nemicham - stejne tema (napr. hodnoty/svetonazor) zni v 155 pozitivne, v 156 negativne.',
 'col 153 (poslech v aplikaci) NENI bug-report: vetsina lidi bud nezkusila, nebo vysvetluje, proc pouziva jinou platformu (agregace vsech podcastu, zvyklost). Skutecne tech. problemy jsou mensina.',
 'col 154: misto vagnich tematu vytazeny jmenovane zdroje + typ zdroje. Nejcastejsi "jiny" zdroj je Denik N.',
 'Strojove citelne kodovani (boolean priznaky na respondenta): respekt_otevrene_kodovano.csv.',
 '',
 'Pocet textovych odpovedi podle otazky:',
 f'  col 154  Zdroje krome Respektu ....... {N154}',
 f"  col 155  V cem vyjimecny ............. {results['155']['N']}",
 f"  col 156  Co vadi / zlepsit ........... {results['156']['N']}",
 f"  col 153  Poslech v aplikaci .......... {results['153']['N']}",
]
r = 4
for ln in notes:
    ws.cell(row=r, column=1, value=ln).font = F(size=10, bold=ln.endswith(':'))
    r += 1
ws.column_dimensions['A'].width = 120

# --- helper pro tematicky list ---
def theme_sheet(name, qkey, title, sub, extra_col=None):
    ws = wb.create_sheet(name)
    title_block(ws, title, sub)
    Nt = results[qkey]['N']
    ws['A4'] = 'n textovych odpovedi:'; ws['A4'].font = F(bold=True, size=10)
    ws['B4'] = Nt; ws['B4'].font = F(size=10)
    hdr = 6
    cols = ['Tema', 'n', '% z text.', 'Priklad 1', 'Priklad 2']
    for j, h in enumerate(cols, 1):
        ws.cell(row=hdr, column=j, value=h)
    style_header(ws, hdr, len(cols))
    rr = hdr + 1
    for row in results[qkey]['rows']:
        ws.cell(row=rr, column=1, value=row['tema']).font = F(size=10, bold=True)
        ws.cell(row=rr, column=2, value=row['n']).font = F(size=10)
        pc = ws.cell(row=rr, column=3, value=f'=B{rr}/$B$4'); pc.font = F(size=10); pc.number_format = '0.0%'
        ex = row['ex'] + ['', '']
        ws.cell(row=rr, column=4, value=ex[0]).font = F(size=9, italic=True, color='494E56')
        ws.cell(row=rr, column=5, value=ex[1]).font = F(size=9, italic=True, color='494E56')
        for j in range(1, 6):
            ws.cell(row=rr, column=j).alignment = Alignment(vertical='top', wrap_text=(j>=4))
            ws.cell(row=rr, column=j).border = BORD
        rr += 1
    for col, w in zip('ABCDE', [34, 7, 10, 60, 60]):
        ws.column_dimensions[col].width = w

theme_sheet('155_Chvala', '155', 'V cem je Respekt vyjimecny (chvala)',
            'Pozitivni temata. Multi-label. Hodnoty/svetonazor se objevuji POZITIVNE (souzneni se ctenarstvem).')
theme_sheet('156_Vytky', '156', 'Co vadi a co zlepsit (vytky / namety)',
            'Negativni/konstruktivni. "Audio/AI hlas" ma smisenou valenci (cast chvali AI nacteni, cast ho kritizuje).')
theme_sheet('153_Poslech', '153', 'Poslech v aplikaci',
            'REFRAME: prevazne "proc jinou platformu" + spokojenost; technicke problemy jsou mensina.')

# --- 154 Zdroje ---
ws = wb.create_sheet('154_Zdroje')
title_block(ws, 'Informacni zdroje krome Respektu', 'Jmenovane zdroje (entity) + typ. Multi-label (az 3 zdroje na osobu). Nejcastejsi: Denik N.')
ws['A4'] = 'n textovych odpovedi:'; ws['A4'].font = F(bold=True, size=10)
ws['B4'] = N154; ws['B4'].font = F(size=10)
# tabulka zdroju
hdr = 6
for j, h in enumerate(['Zdroj', 'Typ', 'n', '% z text.'], 1):
    ws.cell(row=hdr, column=j, value=h)
style_header(ws, hdr, 4)
rr = hdr + 1
for row in out_rows:
    ws.cell(row=rr, column=1, value=row['zdroj']).font = F(size=10, bold=True)
    ws.cell(row=rr, column=2, value=row['typ']).font = F(size=10, color='494E56')
    ws.cell(row=rr, column=3, value=row['n']).font = F(size=10)
    pc = ws.cell(row=rr, column=4, value=f'=C{rr}/$B$4'); pc.font = F(size=10); pc.number_format = '0.0%'
    for j in range(1, 5):
        ws.cell(row=rr, column=j).border = BORD
    rr += 1
# souhrn typu (vedle, sloupce F-H)
ws.cell(row=hdr, column=6, value='Typ zdroje (souhrn)')
ws.cell(row=hdr, column=7, value='n')
ws.cell(row=hdr, column=8, value='% z text.')
style_header(ws, hdr, 8)
tr = hdr + 1
for typ, n in type_rows:
    ws.cell(row=tr, column=6, value=typ).font = F(size=10, bold=True)
    ws.cell(row=tr, column=7, value=n).font = F(size=10)
    pc = ws.cell(row=tr, column=8, value=f'=G{tr}/$B$4'); pc.font = F(size=10); pc.number_format = '0.0%'
    for j in (6, 7, 8):
        ws.cell(row=tr, column=j).border = BORD
    tr += 1
for col, w in zip('ABCDEFGH', [44, 26, 7, 10, 3, 30, 7, 10]):
    ws.column_dimensions[col].width = w

wb.save(OUT_DIR / 'respekt_otevrene_taxonomie.xlsx')
print('Saved xlsx + coded CSV.')
print('coded shape:', coded.shape)
for q in ['155','156','153']:
    print(f"\n{q} (N={results[q]['N']}):")
    for row in results[q]['rows']:
        print(f"   {row['n']:5d} | {row['tema']}")
print(f"\n154 zdroje (N={N154}) - top:")
for row in out_rows[:12]:
    print(f"   {row['n']:5d} | {row['zdroj']}")
print("154 typy:")
for typ,n in type_rows:
    print(f"   {n:5d} | {typ}")
